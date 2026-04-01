"""
engine/excel_handler.py

核心变更（Phase 2 Task 6a）：
- .xlsx 文件使用 openpyxl read_only=True 逐行迭代，避免一次性全量加载进内存。
- .xls 保持 xlrd 路径（xlrd 不支持 read_only，但 .xls 已是遗留格式）。
- 空 Sheet / 加密 Excel 异常捕获，记录 WARNING 后跳过。
"""
from pathlib import Path
from typing import Generator
import datetime

import pandas as pd
from loguru import logger

from engine.matcher import MatchRule, match_row


def _normalize_cell(value) -> str:
    """
    将单元格值统一转为人类可读字符串：
    - datetime / Timestamp → 格式化为 YYYY/M/D（与 Excel 显示习惯一致）
    - float 且值为整数（如 123.0）→ 转为 "123"
    - NaN / None → 空字符串
    - 其余类型 → str()

    Examples:
        >>> import pandas as pd, datetime
        >>> _normalize_cell(datetime.datetime(2026, 3, 31))
        '2026/3/31'
        >>> _normalize_cell(pd.Timestamp('2026-03-31'))
        '2026/3/31'
        >>> _normalize_cell(123.0)
        '123'
        >>> _normalize_cell(float('nan'))
        ''
        >>> _normalize_cell(None)
        ''
        >>> _normalize_cell('张三')
        '张三'
    """
    if value is None:
        return ""

    # pandas / Python float NaN
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    # datetime 对象（Python 标准库 or pandas Timestamp）
    if isinstance(value, (datetime.datetime, pd.Timestamp)):
        return f"{value.year}/{value.month}/{value.day}"

    # date 对象（不含 time）
    if isinstance(value, datetime.date):
        return f"{value.year}/{value.month}/{value.day}"

    # 浮点数：如果实际是整数则去掉 .0
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)

    return str(value)


def _row_to_str(row: pd.Series) -> str:
    """
    将一行中每个单元格经 _normalize_cell() 处理后，
    用空格拼接为单一字符串，供 matcher 使用。
    """
    return " ".join(_normalize_cell(v) for v in row)


def _row_values_to_str(values) -> str:
    """
    将一个值列表（来自 openpyxl 的 row.value tuple）经
    _normalize_cell() 处理后，用空格拼接为单一字符串。
    """
    return " ".join(_normalize_cell(v) for v in values)


def _scan_xlsx_readonly(
    filepath: Path,
    rule: MatchRule,
    source_name: str,
) -> Generator[dict, None, None]:
    """
    使用 openpyxl read_only=True 逐行扫描 .xlsx 文件（Task 6a）。
    内存占用仅与单行数据量相关，与文件总大小无关。
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"无法打开 .xlsx（可能已加密或损坏）[{source_name}]: {e}")
        return

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter, None)
        except StopIteration:
            logger.warning(f"  跳过空 Sheet: {source_name} / {sheet_name}")
            continue

        if header_row is None or all(v is None for v in header_row):
            logger.warning(f"  跳过空 Sheet: {source_name} / {sheet_name}")
            continue

        # 核心修复：去除 openpyxl 读取出的尾部幽灵空列（导致总列数超过16384）
        header_list = list(header_row)
        while header_list and header_list[-1] is None:
            header_list.pop()

        # 核心修复2：表头清理（去除换行、空格），使得“班 级”和“班级”能够被合并在同一直列
        headers = []
        for i, h in enumerate(header_list):
            if h is not None:
                # 统一转字符串，消灭隐形空格和回车
                clean_h = str(h).replace(" ", "").replace("　", "").replace("\n", "").replace("\r", "")
                headers.append(clean_h if clean_h else f"col_{i}")
            else:
                headers.append(f"col_{i}")

        row_count = 0
        empty_lines_streak = 0
        
        for raw_row in rows_iter:
            row_count += 1
            
            # 性能防御：如果 openpyxl 返回了数万行空行（表格格式异常所致），连续 50 行空则强行中断
            if all(v is None for v in raw_row):
                empty_lines_streak += 1
                if empty_lines_streak > 50:
                    break
                continue
            else:
                empty_lines_streak = 0
                
            values = list(raw_row)
            # 丢弃超过有效表头宽度的垃圾单元格数据
            if len(values) > len(headers):
                values = values[:len(headers)]
                
            row_str = _row_values_to_str(values)
            # 添加日志：打印归一化后即将用于匹配的字符串
            # logger.debug(f"[{source_name} - {sheet_name}] row_str: {row_str}")
            
            if match_row(row_str, rule):
                row_dict = dict(zip(headers, values))
                yield {
                    "source_file":  source_name,
                    "source_sheet": sheet_name,
                    **row_dict,
                }

        if row_count == 0:
            logger.warning(f"  Sheet 无数据行: {source_name} / {sheet_name}")

    wb.close()


def _scan_xls_pandas(
    filepath: Path,
    rule: MatchRule,
    source_name: str,
) -> Generator[dict, None, None]:
    """使用 pandas + xlrd 扫描旧版 .xls 文件（无法使用 read_only 模式）。"""
    try:
        xl = pd.ExcelFile(filepath, engine='xlrd')
    except Exception as e:
        logger.warning(f"无法打开 .xls（可能已加密或损坏）[{source_name}]: {e}")
        return

    for sheet_name in xl.sheet_names:
        try:
            df = xl.parse(sheet_name, header=0)
        except Exception as e:
            logger.warning(f"  跳过 Sheet [{sheet_name}] 无法读取: {e}")
            continue

        if df.empty:
            logger.warning(f"  跳过空 Sheet: {source_name} / {sheet_name}")
            continue

        # 对 xls 数据表的表头做相同规格的清洗去空处理
        new_cols = []
        for c in df.columns:
            c_str = str(c)
            if "Unnamed" in c_str:
                new_cols.append(c_str)
            else:
                clean_c = c_str.replace(" ", "").replace("　", "").replace("\n", "").replace("\r", "")
                new_cols.append(clean_c if clean_c else c_str)
        df.columns = new_cols

        for _, row in df.iterrows():
            row_str = _row_to_str(row)
            # 添加日志：打印归一化后即将用于匹配的字符串
            logger.debug(f"[{source_name} - {sheet_name}] row_str: {row_str}")
            
            if match_row(row_str, rule):
                yield {
                    "source_file":  source_name,
                    "source_sheet": sheet_name,
                    **row.to_dict(),
                }


def scan_excel(
    filepath: Path,
    rule: MatchRule
) -> Generator[dict, None, None]:
    """
    遍历每个 Sheet，对每行调用 _normalize_cell() / _row_to_str() 后用 rule 匹配。

    yield 格式：
    {
      "source_file":  str,   # 原始文件名（不含临时路径前缀）
      "source_sheet": str,   # Sheet 名
      **row_data              # 该行所有列的原始值（未经 normalize，保留原始）
    }

    .xlsx → openpyxl read_only=True（节约内存）
    .xls  → xlrd（遗留格式回退）
    加密 Excel / 空 Sheet 记录 WARNING 后跳过，不抛异常。
    """
    source_name = filepath.name
    suffix = filepath.suffix.lower()

    if suffix == '.xlsx':
        yield from _scan_xlsx_readonly(filepath, rule, source_name)
    elif suffix == '.xls':
        yield from _scan_xls_pandas(filepath, rule, source_name)
    else:
        logger.warning(f"不支持的文件格式，跳过: {source_name}")
