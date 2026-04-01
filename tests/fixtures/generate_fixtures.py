"""
生成 tests/fixtures/ 中所需的 fixture zip 文件。

关键难点说明：
    Python zipfile 库在写入含非 ASCII 字符名时，会强制设置 flag_bits |= 0x800（UTF-8 标志）。
    为了伪造一个"真正的 GBK 压缩包"（即 Windows 中文系统压缩工具的产物），
    我们使用 struct 直接构造 ZIP 二进制格式，将 GBK 字节原样写入文件名字段，
    并确保 General Purpose Bit Flag 的 bit 11（UTF-8 标志）为 0。
"""
import os
import struct
import zlib
import time


def _dos_time(t=None):
    """将 time.struct_time 转换为 DOS time/date 双字。"""
    if t is None:
        t = time.localtime()
    dos_time = (t.tm_hour << 11) | (t.tm_min << 5) | (t.tm_sec // 2)
    dos_date = ((t.tm_year - 1980) << 9) | (t.tm_mon << 5) | t.tm_mday
    return dos_time, dos_date


def _make_local_file_header(
    filename_bytes: bytes, data: bytes, crc: int, dos_time: int, dos_date: int
) -> bytes:
    """构造 Local File Header（signature PK\\x03\\x04）。"""
    return struct.pack(
        '<4s2B4HL2L2H',
        b'PK\x03\x04',   # signature
        20, 0,            # version needed, version made by
        0,                # general purpose bit flag = 0 (NO UTF-8, NO encryption)
        0,                # compression method = stored
        dos_time,
        dos_date,
        crc,              # CRC-32
        len(data),        # compressed size
        len(data),        # uncompressed size
        len(filename_bytes),
        0,                # extra field length
    ) + filename_bytes


def _make_central_dir_record(
    filename_bytes: bytes, data: bytes, crc: int,
    dos_time: int, dos_date: int, local_header_offset: int
) -> bytes:
    """构造 Central Directory File Header（signature PK\\x01\\x02）。"""
    return struct.pack(
        '<4s4B4HL2L5H2L',
        b'PK\x01\x02',   # signature
        20, 0,            # version made by
        20, 0,            # version needed
        0,                # general purpose bit flag = 0
        0,                # compression method = stored
        dos_time,
        dos_date,
        crc,
        len(data),        # compressed size
        len(data),        # uncompressed size
        len(filename_bytes),
        0,                # extra field length
        0,                # file comment length
        0,                # disk number start
        0,                # internal file attributes
        0x20,             # external file attributes (archive)
        local_header_offset,
    ) + filename_bytes


def _make_eocd(num_entries: int, central_dir_size: int, central_dir_offset: int) -> bytes:
    """构造 End of Central Directory Record（signature PK\\x05\\x06）。"""
    return struct.pack(
        '<4s4H2LH',
        b'PK\x05\x06',
        0,                # disk number
        0,                # disk with start of central directory
        num_entries,      # entries on this disk
        num_entries,      # total entries
        central_dir_size,
        central_dir_offset,
        0,                # comment length
    )


def _build_zip(entries: list) -> bytes:
    """
    从 (filename_bytes, file_data) 元组列表构建 ZIP 二进制数据。
    filename_bytes 直接原样写入 header，不做任何编码转换。
    所有 flag_bits = 0（无 UTF-8 标志）。
    """
    local_headers = b''
    central_dirs = b''
    offsets = []

    dos_time, dos_date = _dos_time()

    for fname_bytes, data in entries:
        offsets.append(len(local_headers))
        crc = zlib.crc32(data) & 0xFFFFFFFF
        local_headers += _make_local_file_header(fname_bytes, data, crc, dos_time, dos_date)
        local_headers += data

    cd_offset = len(local_headers)
    for i, (fname_bytes, data) in enumerate(entries):
        crc = zlib.crc32(data) & 0xFFFFFFFF
        central_dirs += _make_central_dir_record(
            fname_bytes, data, crc, dos_time, dos_date, offsets[i]
        )

    eocd = _make_eocd(len(entries), len(central_dirs), cd_offset)
    return local_headers + central_dirs + eocd


def create_gbk_zip() -> None:
    """创建 gbk_named.zip：一个 ASCII 文件名 + 一个 GBK 编码文件名的压缩包。"""
    fixture_dir = os.path.dirname(os.path.abspath(__file__))
    zip_path = os.path.join(fixture_dir, 'gbk_named.zip')

    entries = [
        (b'ascii_file.xlsx', b'dummy content 1'),
        # GBK bytes written verbatim — no Python string encoding involved
        ('测试.xlsx'.encode('gbk'), b'dummy content 2'),
    ]

    zip_data = _build_zip(entries)
    with open(zip_path, 'wb') as f:
        f.write(zip_data)

    print(f"GBK fixture written to: {zip_path}")


def create_nested_zip() -> None:
    """
    创建 nested.zip：2 层嵌套 zip，结构如下
        nested.zip
        └── inner.zip
            └── date_data.xlsx   (含日期列和目标关键词的真实 xlsx)
    """
    import io
    import zipfile as stdlib_zipfile
    import openpyxl
    import datetime

    fixture_dir = os.path.dirname(os.path.abspath(__file__))

    # 1. 生成真实的 xlsx（含日期列、目标行、非目标行）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩单"
    ws.append(["姓名", "成绩", "日期", "备注"])
    ws.append(["张三", 95,  datetime.datetime(2026, 3, 31), "优秀"])
    ws.append(["李四", 82,  datetime.datetime(2026, 1, 5),  "良好"])
    ws.append(["王五", 77,  datetime.datetime(2025, 12, 1), "合格"])

    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_bytes = xlsx_buf.getvalue()

    # 2. 构建 inner.zip（用标准 zipfile，文件名 ASCII）
    inner_buf = io.BytesIO()
    with stdlib_zipfile.ZipFile(inner_buf, 'w', compression=stdlib_zipfile.ZIP_STORED) as inner_zf:
        inner_zf.writestr("date_data.xlsx", xlsx_bytes)
    inner_bytes = inner_buf.getvalue()

    # 3. 构建 nested.zip（包含 inner.zip）
    nested_path = os.path.join(fixture_dir, 'nested.zip')
    with stdlib_zipfile.ZipFile(nested_path, 'w', compression=stdlib_zipfile.ZIP_STORED) as outer_zf:
        outer_zf.writestr("inner.zip", inner_bytes)

    print(f"Nested fixture written to: {nested_path}")


def create_corrupted_zip() -> None:
    """创建 corrupted.zip：故意损坏的 ZIP 文件（前缀有效签名但 body 为乱码）。"""
    fixture_dir = os.path.dirname(os.path.abspath(__file__))
    zip_path = os.path.join(fixture_dir, 'corrupted.zip')
    with open(zip_path, 'wb') as f:
        # PK 签名 OK，但数据完全无效
        f.write(b'PK\x03\x04' + b'\xff' * 200)
    print(f"Corrupted fixture written to: {zip_path}")


def create_wide_columns_xlsx() -> None:
    """
    创建 wide_columns.xlsx：3 个 Sheet，每个 Sheet 列名各异，
    用于测试 COMMON 列模式（仅保留所有 Sheet 共有列）。
    """
    import openpyxl
    fixture_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(fixture_dir, 'wide_columns.xlsx')

    wb = openpyxl.Workbook()

    # Sheet1: 姓名, 学号, 成绩, 班级
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["姓名", "学号", "成绩", "班级"])
    ws1.append(["张三", "2101", 95, "A班"])
    ws1.append(["李四", "2102", 82, "B班"])

    # Sheet2: 姓名, 学号, 排名（无"成绩"列）
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["姓名", "学号", "排名"])
    ws2.append(["王五", "2103", 5])

    # Sheet3: 姓名, 学号, 备注（同样无"成绩"）
    ws3 = wb.create_sheet("Sheet3")
    ws3.append(["姓名", "学号", "备注"])
    ws3.append(["赵六", "2104", "优秀"])

    wb.save(xlsx_path)
    print(f"Wide columns fixture written to: {xlsx_path}")


if __name__ == '__main__':
    create_gbk_zip()
    create_nested_zip()
    create_corrupted_zip()
    create_wide_columns_xlsx()

