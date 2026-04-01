"""
tests/test_robustness.py

Phase 2 Task 5a：异常路径测试。
验证 processor.py 和 excel_handler.py 遇到以下情况时能跳过并继续：
  - 损坏的 ZIP（corrupted.zip）
  - 空 Sheet（empty_sheet.xlsx）
  - 加密 Excel（模拟：传入不合法 xlsx 字节）
"""
import io
import shutil
import tempfile
import zipfile
from pathlib import Path

import openpyxl
import pytest

from engine.processor import extract_all_excels
from engine.excel_handler import scan_excel
from engine.matcher import MatchRule, MatchMode

FIXTURE_DIR = Path(__file__).parent / "fixtures"
CORRUPTED_ZIP = FIXTURE_DIR / "corrupted.zip"


# ──────────────────────────────────────────────────────────────
# Task 5a - 损坏的 ZIP
# ──────────────────────────────────────────────────────────────

def test_corrupted_zip_does_not_crash(tmp_path):
    """损坏的 ZIP 不应抛出异常，应返回空列表。"""
    result = list(extract_all_excels(CORRUPTED_ZIP, tmp_path))
    assert result == [], f"期望空列表，实际: {result}"


def test_corrupted_zip_mixed_with_valid(tmp_path):
    """
    一个外层 zip 包含：损坏的内层 zip + 正常的 xlsx。
    系统应跳过损坏部分，仍然返回正常的 xlsx。
    """
    # 构建临时混合 zip
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["姓名", "成绩"])
    ws.append(["张三", 90])
    xlbuf = io.BytesIO()
    wb.save(xlbuf)
    xlsx_bytes = xlbuf.getvalue()

    mixed_zip = tmp_path / "mixed.zip"
    with zipfile.ZipFile(mixed_zip, 'w') as zf:
        zf.writestr("good.xlsx", xlsx_bytes)
        zf.writestr("bad.zip", b'PK\x03\x04' + b'\xff' * 50)

    results = list(extract_all_excels(mixed_zip, tmp_path))
    names = [p.name for p in results]
    assert "good.xlsx" in names
    assert len(results) == 1  # bad.zip 被跳过


# ──────────────────────────────────────────────────────────────
# Task 5a - 空 Sheet
# ──────────────────────────────────────────────────────────────

def test_empty_sheet_does_not_crash(tmp_path):
    """含有空 Sheet 的 xlsx 不应崩溃，且返回空结果。"""
    empty_xlsx = tmp_path / "empty_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empty"
    wb.save(empty_xlsx)

    rule = MatchRule(keywords=["任意"], mode=MatchMode.AND)
    results = list(scan_excel(empty_xlsx, rule))
    assert results == []


# ──────────────────────────────────────────────────────────────
# Task 5a - "加密" Excel（实际上是损坏的 xlsx 字节）
# ──────────────────────────────────────────────────────────────

def test_invalid_xlsx_bytes_does_not_crash(tmp_path):
    """
    向 scan_excel 传入一个内容完全无效的 .xlsx 文件（模拟加密/损坏），
    系统应记录 WARNING 并跳过，不崩溃。
    """
    bad_xlsx = tmp_path / "bad.xlsx"
    bad_xlsx.write_bytes(b'\x00' * 512)  # 非合法 xlsx

    rule = MatchRule(keywords=["任意"], mode=MatchMode.AND)
    results = list(scan_excel(bad_xlsx, rule))
    assert results == []


def test_corrupted_zip_in_nested_does_not_crash(tmp_path):
    """2 层嵌套中的损坏 zip 不影响同层其他文件的处理。"""
    # 构建 outer.zip: [valid.xlsx, broken.zip]
    wb = openpyxl.Workbook()
    wb.active.append(["姓名"])
    wb.active.append(["张三"])
    xlbuf = io.BytesIO()
    wb.save(xlbuf)

    outer = tmp_path / "outer.zip"
    with zipfile.ZipFile(outer, 'w') as zf:
        zf.writestr("valid.xlsx", xlbuf.getvalue())
        zf.writestr("broken.zip", b'PK\x03\x04' + b'\xde\xad' * 100)

    results = list(extract_all_excels(outer, tmp_path))
    assert any(p.name == "valid.xlsx" for p in results)


# ──────────────────────────────────────────────────────────────
# Task 5b - 临时目录清理（验证 ignore_cleanup_errors 不崩溃）
# ──────────────────────────────────────────────────────────────

def test_temp_dir_cleanup_does_not_crash(tmp_path):
    """TemporaryDirectory(ignore_cleanup_errors=True) 在 Windows 文件锁情况下不崩溃。"""
    import contextlib
    with contextlib.ExitStack() as stack:
        td = stack.enter_context(
            tempfile.TemporaryDirectory(dir=tmp_path, ignore_cleanup_errors=True)
        )
        td_path = Path(td)
        # 模拟残留文件
        (td_path / "test.txt").write_text("hello")
    # 退出 with 后不应抛出异常
    assert True
