"""
tests/test_aggregator.py

Phase 2 Task 6a：测试 aggregate() 的正常模式与流式写入模式。
"""
import pytest
import tempfile
import shutil
from pathlib import Path
import openpyxl

from engine.aggregator import aggregate, STREAM_THRESHOLD


def _make_records(n: int, prefix: str = "file.xlsx") -> list:
    """生成 n 条模拟匹配记录。"""
    return [
        {
            "source_file": prefix,
            "source_sheet": "Sheet1",
            "姓名": f"学生{i}",
            "成绩": i % 100,
        }
        for i in range(n)
    ]


# ──────────────────────────────────────────────────────────────
# 正常模式（≤ stream_threshold）
# ──────────────────────────────────────────────────────────────

def test_aggregate_normal_mode(tmp_path):
    records = _make_records(10)
    output = tmp_path / "out.xlsx"

    df, total = aggregate(iter(records), output, stream_threshold=STREAM_THRESHOLD)

    assert df is not None
    assert total == 10
    assert list(df.columns[:2]) == ["source_file", "source_sheet"]
    assert len(df) == 10


def test_aggregate_meta_cols_at_left(tmp_path):
    """确认 source_file / source_sheet 始终在最左侧，与原始 dict 顺序无关。"""
    records = [
        {"姓名": "张三", "source_file": "a.xlsx", "source_sheet": "S1", "成绩": 90}
    ]
    df, _ = aggregate(iter(records), tmp_path / "out.xlsx")
    assert df.columns[0] == "source_file"
    assert df.columns[1] == "source_sheet"


def test_aggregate_empty_records(tmp_path):
    df, total = aggregate(iter([]), tmp_path / "out.xlsx")
    assert total == 0
    assert df is not None and df.empty


def test_aggregate_outer_join(tmp_path):
    """来自不同来源的列用 outer join 对齐，缺失值为 NaN。"""
    records = [
        {"source_file": "a.xlsx", "source_sheet": "S1", "姓名": "张三", "成绩": 90},
        {"source_file": "b.xlsx", "source_sheet": "S1", "姓名": "李四", "备注": "良好"},
    ]
    df, _ = aggregate(iter(records), tmp_path / "out.xlsx")
    assert "成绩" in df.columns
    assert "备注" in df.columns
    assert df.loc[0, "备注"] != df.loc[0, "备注"]  # NaN != NaN


# ──────────────────────────────────────────────────────────────
# 流式模式（> stream_threshold）
# ──────────────────────────────────────────────────────────────

def test_aggregate_stream_mode_triggers(tmp_path):
    """当记录数超过阈值时，应切换流式模式返回 (None, n)。"""
    small_threshold = 5
    records = _make_records(10)
    output = tmp_path / "stream_out.xlsx"

    df, total = aggregate(iter(records), output, stream_threshold=small_threshold)

    assert df is None, "流式模式应返回 None DataFrame"
    assert total == 10
    assert output.exists(), "流式模式应直接写出文件"


def test_aggregate_stream_output_is_valid_xlsx(tmp_path):
    """流式写出的文件必须是有效的 xlsx，且包含正确的行列数。"""
    small_threshold = 3
    n = 10
    records = _make_records(n)
    output = tmp_path / "stream_out.xlsx"

    df, total = aggregate(iter(records), output, stream_threshold=small_threshold)

    assert output.exists()
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # 第一行是 header
    assert rows[0][0] == "source_file"
    # 数据行数量
    assert len(rows) - 1 == n


def test_aggregate_stream_meta_cols_at_left(tmp_path):
    """流式模式下 source_file / source_sheet 也应在最左侧。"""
    small_threshold = 2
    records = [
        {"姓名": "张三", "source_file": "a.xlsx", "source_sheet": "S1"},
        {"姓名": "李四", "source_file": "a.xlsx", "source_sheet": "S1"},
        {"姓名": "王五", "source_file": "a.xlsx", "source_sheet": "S1"},
    ]
    output = tmp_path / "stream_out.xlsx"
    aggregate(iter(records), output, stream_threshold=2)

    wb = openpyxl.load_workbook(output)
    header = [c.value for c in next(wb.active.iter_rows())]
    assert header[0] == "source_file"
    assert header[1] == "source_sheet"
